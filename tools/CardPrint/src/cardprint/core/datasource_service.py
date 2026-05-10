from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .errors import CardPrintError
from .models import DEFAULT_HEADER_ALIASES


def normalize_row_keys(
    row: dict[str, Any],
    aliases: dict[str, str] | None = None,
) -> dict[str, Any]:
    alias_map = aliases or DEFAULT_HEADER_ALIASES
    normalized: dict[str, Any] = {}
    for raw_key, value in row.items():
        if raw_key is None:
            continue
        key = str(raw_key).strip()
        mapped_key = alias_map.get(key.lower(), key)
        normalized[mapped_key] = value
    return normalized


def load_rows_from_csv(path: str | Path, aliases: dict[str, str] | None = None) -> list[dict[str, Any]]:
    csv_path = Path(path)
    if not csv_path.exists():
        raise CardPrintError(
            code="CSV_NOT_FOUND",
            message="CSV 文件不存在。",
            details={"path": str(path)},
        )
    rows: list[dict[str, Any]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(normalize_row_keys(dict(row), aliases))
    return rows


def load_rows_from_xlsx(path: str | Path, aliases: dict[str, str] | None = None) -> list[dict[str, Any]]:
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        raise CardPrintError(
            code="XLSX_NOT_FOUND",
            message="XLSX 文件不存在。",
            details={"path": str(path)},
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CardPrintError(
            code="XLSX_DEP_MISSING",
            message="读取 XLSX 需要安装 openpyxl。",
            details={"hint": "pip install openpyxl"},
        ) from exc

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        return []
    normalized_headers = [str(h).strip() if h is not None else "" for h in headers]
    rows: list[dict[str, Any]] = []
    for row_values in rows_iter:
        raw_row: dict[str, Any] = {}
        for idx, value in enumerate(row_values):
            header = normalized_headers[idx] if idx < len(normalized_headers) else f"col_{idx}"
            if header:
                raw_row[header] = value
        rows.append(normalize_row_keys(raw_row, aliases))
    return rows


def load_rows(path: str | Path, aliases: dict[str, str] | None = None) -> list[dict[str, Any]]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return load_rows_from_csv(path, aliases=aliases)
    if ext in {".xlsx", ".xlsm"}:
        return load_rows_from_xlsx(path, aliases=aliases)
    raise CardPrintError(
        code="UNSUPPORTED_DATA_FILE",
        message="仅支持 CSV/XLSX 文件导入。",
        details={"path": str(path), "supported": [".csv", ".xlsx", ".xlsm"]},
    )
